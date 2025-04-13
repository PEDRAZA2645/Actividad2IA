from collections import deque
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import accuracy_score
from openpyxl import Workbook
from openpyxl.styles import Border, Side


# Clase Estacion
class Estacion:
    def __init__(self, nombre):
        self.nombre = nombre
        self.conexiones = {}

    def agregar_conexion(self, linea, estacion_destino, tiempo_viaje):
        self.conexiones[linea] = (estacion_destino, tiempo_viaje)

    def __repr__(self):
        return self.nombre


# Clase SistemaBasadoEnReglas
class SistemaBasadoEnReglas:
    def __init__(self):
        self.reglas = []  # Lista de reglas (condición, acción)
        self.hechos = []  # Lista de hechos iniciales

    def agregar_regla(self, condicion, accion):
        self.reglas.append((condicion, accion))

    def agregar_hecho(self, hecho):
        self.hechos.append(hecho)

    def ejecutar(self):
        cambios = True
        while cambios:
            cambios = False
            for condicion, accion in self.reglas:
                if condicion(self.hechos):
                    nuevos_hechos = accion(self.hechos)
                    for hecho in nuevos_hechos:
                        if hecho not in self.hechos:
                            self.hechos.append(hecho)
                            cambios = True


# Función para generar el sistema de transporte y exportar datos
def sistema_transporte(estaciones, archivo_excel="rutas_generadas.xlsx", archivo_csv="rutas_generadas.csv"):
    sistema = SistemaBasadoEnReglas()

    # Base de conocimientos: conexiones entre estaciones
    conexiones = []
    for estacion in estaciones.values():
        for linea, (destino, tiempo) in estacion.conexiones.items():
            conexiones.append((estacion.nombre, destino.nombre, linea, tiempo))

    # Regla 1: Conexión directa
    def regla_conexion_directa(hechos):
        for origen, destino, linea, tiempo in conexiones:
            if (origen, destino) not in [(h[0], h[1]) for h in hechos]:
                return True
        return False

    def accion_conexion_directa(hechos):
        nuevos_hechos = []
        for origen, destino, linea, tiempo in conexiones:
            if (origen, destino) not in [(h[0], h[1]) for h in hechos]:
                nuevos_hechos.append((origen, destino, linea, tiempo))
        return nuevos_hechos

    sistema.agregar_regla(regla_conexion_directa, accion_conexion_directa)

    # Regla 2: Conexión indirecta
    def regla_conexion_indirecta(hechos):
        for h1 in hechos:
            for h2 in hechos:
                if h1[1] == h2[0] and (h1[0], h2[1]) not in [(h[0], h[1]) for h in hechos]:
                    if h1[3] + h2[3] <= 120:
                        return True
        return False

    def accion_conexion_indirecta(hechos):
        nuevos_hechos = []
        for h1 in hechos:
            for h2 in hechos:
                if h1[1] == h2[0] and (h1[0], h2[1]) not in [(h[0], h[1]) for h in hechos]:
                    tiempo_total = h1[3] + h2[3]
                    if tiempo_total <= 120:
                        nuevos_hechos.append((h1[0], h2[1], f"{h1[2]} + {h2[2]}", tiempo_total))
        return nuevos_hechos

    sistema.agregar_regla(regla_conexion_indirecta, accion_conexion_indirecta)

    # Hechos iniciales
    for origen, destino, linea, tiempo in conexiones:
        sistema.agregar_hecho((origen, destino, linea, tiempo))

    # Ejecutar el sistema
    sistema.ejecutar()

    # Crear un DataFrame con los resultados
    datos_rutas = []
    for hecho in sistema.hechos:
        es_optima = "Si" if hecho[3] <= 60 else "No"
        datos_rutas.append([hecho[0], hecho[1], hecho[2], hecho[3], es_optima])

    columnas = ["Estacion Origen", "Estacion Destino", "Linea", "Tiempo (min)", "Ruta Optima"]
    df_rutas = pd.DataFrame(datos_rutas, columns=columnas)

    # Exportar a Excel con bordes
    with pd.ExcelWriter(archivo_excel, engine="openpyxl") as writer:
        df_rutas.to_excel(writer, index=False, sheet_name="Rutas")
        workbook = writer.book
        worksheet = writer.sheets["Rutas"]

        # Definir el estilo de bordes
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Aplicar bordes a todas las celdas
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

    print(f"Rutas exportadas a {archivo_excel} con bordes en el grid.")

    # Exportar a CSV
    df_rutas.to_csv(archivo_csv, index=False, encoding="utf-8")
    print(f"Rutas exportadas a {archivo_csv}.")


# Clase GestorDatos
class GestorDatos:
    def __init__(self, archivo_datos):
        self.archivo_datos = archivo_datos
        self.datos = None
        self.X = None
        self.y = None

    def cargar_datos(self):
        """Carga los datos desde un archivo CSV con codificación utf-8."""
        self.datos = pd.read_csv(self.archivo_datos, encoding='utf-8')
        print("Datos cargados con éxito.\n", self.datos.head())

    def preparar_datos(self):
        """Prepara los datos para el modelo supervisado."""
        le = LabelEncoder()
        self.datos['Estacion Origen'] = le.fit_transform(self.datos['Estacion Origen'])
        self.datos['Estacion Destino'] = le.fit_transform(self.datos['Estacion Destino'])
        self.datos['Linea'] = le.fit_transform(self.datos['Linea'])

        self.X = self.datos.drop("Ruta Optima", axis=1)
        self.y = self.datos["Ruta Optima"]
        return self.X, self.y


# Clase ModeloAprendizaje
class ModeloAprendizaje:
    def __init__(self):
        self.modelo = DecisionTreeClassifier()

    def dividir_datos(self, X, y):
        self.X_train, self.X_test, self.y_train, self.y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    def entrenar_modelo(self):
        self.modelo.fit(self.X_train, self.y_train)

    def evaluar_modelo(self):
        y_pred = self.modelo.predict(self.X_test)
        print(f"Precisión: {accuracy_score(self.y_test, y_pred)}")

    def predecir(self, ejemplo):
        prediccion = self.modelo.predict(ejemplo)
        print(f"Predicción para el ejemplo: {prediccion[0]}")
        return prediccion[0]


# Main
if __name__ == "__main__":
    # Generar datos con el sistema basado en reglas
    estaciones_data = {
        "Tatooine": Estacion("Tatooine"),
        "Alderaan": Estacion("Alderaan"),
        "Yavin IV": Estacion("Yavin IV"),
        "Hoth": Estacion("Hoth"),
        "Dagobah": Estacion("Dagobah"),
        "Bespin": Estacion("Bespin"),
        "Endor": Estacion("Endor"),
        "Naboo": Estacion("Naboo"),
        "Coruscant": Estacion("Coruscant"),
        "Kamino": Estacion("Kamino"),
        "Mandalore": Estacion("Mandalore"),
    }

    # Agregar todas las conexiones especificadas
    estaciones_data["Tatooine"].agregar_conexion("Ruta 1", estaciones_data["Alderaan"], 10)
    estaciones_data["Alderaan"].agregar_conexion("Ruta 2", estaciones_data["Yavin IV"], 20)
    estaciones_data["Yavin IV"].agregar_conexion("Ruta 3", estaciones_data["Hoth"], 30)
    estaciones_data["Hoth"].agregar_conexion("Ruta 4", estaciones_data["Dagobah"], 40)
    estaciones_data["Dagobah"].agregar_conexion("Ruta 5", estaciones_data["Bespin"], 50)
    estaciones_data["Bespin"].agregar_conexion("Ruta 6", estaciones_data["Endor"], 60)
    estaciones_data["Endor"].agregar_conexion("Ruta 7", estaciones_data["Naboo"], 70)
    estaciones_data["Naboo"].agregar_conexion("Ruta 8", estaciones_data["Coruscant"], 80)
    estaciones_data["Coruscant"].agregar_conexion("Ruta 9", estaciones_data["Kamino"], 90)
    estaciones_data["Kamino"].agregar_conexion("Ruta 10", estaciones_data["Mandalore"], 100)

    # Generar y exportar el sistema de transporte
    sistema_transporte(estaciones_data)

    # Entrenar modelo supervisado
    gestor = GestorDatos("rutas_generadas.csv")
    gestor.cargar_datos()
    X, y = gestor.preparar_datos()

    modelo = ModeloAprendizaje()
    modelo.dividir_datos(X, y)
    modelo.entrenar_modelo()
    modelo.evaluar_modelo()

    # Ejemplo de predicción
    ejemplo = X.iloc[0].values.reshape(1, -1)
    modelo.predecir(ejemplo)